import openpyxl
import word_extractor
import sys
# from googletrans import Translator


class OutputExtractedWordToExcel (word_extractor.WordExtractorFromFolder):
    # translator = Translator()

    def __init__(self, src_folder, language, wb_output="", wb_reference=""):
        super().__init__(src_folder, language)
        self.WB_REFERENCE_NAME = wb_reference
        self.WB_REFERENCE = openpyxl.load_workbook(self.WB_REFERENCE_NAME)
        self.WS_REFERENCE = self.WB_REFERENCE.worksheets[0]

        self.WB_OUTPUT_NAME = wb_output
        self.wb_output = openpyxl.load_workbook(self.WB_OUTPUT_NAME)
        self.ws_output = self.wb_output.worksheets[0]
        self.create_word_dictionary_from_folder()
        self.__read_from_reference()

    def __read_from_reference(self):
        self.REFERENCE_WORD_DICT = {}
        REFEREMCE_MAX_ROW = self.WS_REFERENCE.max_row
        for row in range(3, REFEREMCE_MAX_ROW+1):
            term = self.WS_REFERENCE.cell(row=row, column=2).value
            bilingual_translation = self.WS_REFERENCE.cell(
                row=row, column=3).value
            self.REFERENCE_WORD_DICT[term] = bilingual_translation

    def __write_to_wb(self):
        current_row = 2
        for word_information in self.get_word_dictionary_information().values():
            self.ws_output.cell(row=current_row, column=1).value = ""
            self.ws_output.cell(
                row=current_row, column=2).value = word_information["file_name"]
            self.ws_output.cell(
                row=current_row, column=3).value = word_information["line_number"]
            self.ws_output.cell(
                row=current_row, column=4).value = word_information["word"]
            self.ws_output.cell(row=current_row, column=5).value = self.__get_bilingual_translation(
                word=word_information["word"])
            self.ws_output.cell(row=current_row, column=6).value = self.__get_existence(
                word=word_information["word"])
            self.ws_output.cell(row=current_row, column=7).value = self.__get_details(
                word=word_information["word"])
            current_row += 1

    # 英語を和訳して返す
    def __get_bilingual_translation(self, word=""):
        return "日本語"
        # return OutputExtractedWordToExcel.translator.translate(word, src='en', dest='ja')

    # 用語に存在するか（有 or 無）

    def __get_existence(self, word=""):
        if self.__get_existence_on_ja(word) or self.__get_existence_on_en(word):
            return "有"
        return "無"

    # 用語(和訳した結果)が存在
    def __get_existence_on_ja(self, word=""):
        # word_ja = self.__get_bilingual_translation(word=word)
        # if word_ja in self.REFERENCE_WORD_DICT:
        #     return True
        return False

    # 対訳が存在

    def __get_existence_on_en(self, word=""):
        if word in list(self.REFERENCE_WORD_DICT.values()):
            return True
        return False

    # 詳細を返す

    def __get_details(self, word=""):
        text = ""
        written_flg = False
        if not self.__get_existence_on_en(word=word):
            if self.__get_existence_on_ja(word=word):
                text += "対訳が合わない"
                written_flg = True
            else:
                if written_flg:
                    text += "、"
                text += "対訳リストに存在しない"
        return text

    def output_extracted_word_to_excel(self):
        self.__write_to_wb()
        self.WB_REFERENCE.close()
        self.wb_output.save(self.WB_OUTPUT_NAME)
        self.wb_output.close()


if __name__ == "__main__":
    args = sys.argv
    if len(args) != 5:
        print(
            "Please run as --> \"pipenv run python output_to_excel.py [file or folder path] [language] [reference file path] [output file path]\"")
        sys.exit()

    path = args[1]
    language = args[2]
    reference_file_path = args[3]
    output_file_path = args[4]

    if not language in OutputExtractedWordToExcel.CORRESPONDED_LANGUAGE:
        print("Please choose language from \"c\" or \"c++\" or \"python\"")
        sys.exit()

    if path[-1] == "/":
        output_extracted_word_to_excel = OutputExtractedWordToExcel(
            src_folder=path, language=language, wb_reference=reference_file_path, wb_output=output_file_path)
        output_extracted_word_to_excel.output_extracted_word_to_excel()
    else: # TODO 実装
        pass
